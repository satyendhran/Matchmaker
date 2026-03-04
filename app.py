import json
import secrets
from datetime import datetime

import io
import pandas as pd
from flask import Flask, jsonify, make_response, render_template, request, send_file
from flask_cors import CORS
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from plugin_loader import PluginLoader
from tournament_calculators import (
    PercentagePointsCalculator,
    RankingPointsCalculator,
    StandardPointsCalculator,
    ThreePointsCalculator,
)
from tournament_core import (
    MatchmakingStrategyRegistry,
    MatchResult,
    PointsCalculatorRegistry,
    RoundConfig,
    generate_id,
)
from tournament_repository_ext import ExtendedSQLiteTournamentRepository
from tournament_service import TournamentService
from tournament_strategies import (
    FreeForAllStrategy,
    RoundRobinStrategy,
    SingleEliminationStrategy,
    SwissStrategy,
)


from auth_repository import SQLiteUserRepository
from auth_service import AuthenticationService, AuthorizationService, UserManagementService
from auth_middleware import create_auth_middleware
from auth_routes import init_auth_routes
from auth_core import UserRole, UserStatus

try:
    from plugins.chess_fide_swiss import (
        ChessFideDB,
        ChessFidePointsCalculator,
        ChessFideSwissStrategy,
        FideReports,
        FIDE_CATEGORIES,
        FIDE_TITLES,
        TiebreakEngine,
        TitleNormChecker,
    )
    FIDE_AVAILABLE = True
except ImportError:
    FIDE_AVAILABLE = False
    
    class _FideStub:
        """Placeholder when the FIDE plugin is not installed."""
        def __init__(self, *args, **kwargs):
            raise RuntimeError(
                "FIDE plugin (plugins/chess_fide_swiss.py) is not installed. "
                "Please add the plugin to use FIDE features."
            )
    ChessFideDB = ChessFidePointsCalculator = ChessFideSwissStrategy = _FideStub
    FideReports = TiebreakEngine = TitleNormChecker = _FideStub
    FIDE_CATEGORIES = FIDE_TITLES = {}

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
CORS(app, supports_credentials=True)


class TournamentWebApp:
    """Service layer for Flask application with authentication."""

    def __init__(self):
        self.repository = ExtendedSQLiteTournamentRepository()
        self.strategy_registry = MatchmakingStrategyRegistry()
        self.calculator_registry = PointsCalculatorRegistry()

        
        self.user_repository = SQLiteUserRepository()
        self.auth_service = AuthenticationService(self.user_repository)
        self.authz_service = AuthorizationService(self.user_repository)
        self.user_mgmt_service = UserManagementService(
            self.user_repository, self.auth_service
        )
        self.auth_middleware = create_auth_middleware(
            self.auth_service, self.authz_service
        )

        
        if FIDE_AVAILABLE:
            self.fide_db = ChessFideDB()
            self.tiebreak_engine = TiebreakEngine(self.fide_db, self.repository)
            self.fide_reports = FideReports(
                self.fide_db, self.tiebreak_engine, self.repository
            )
            self.title_norm_checker = TitleNormChecker(
                self.fide_db, self.tiebreak_engine
            )
        else:
            self.fide_db = None
            self.tiebreak_engine = None
            self.fide_reports = None
            self.title_norm_checker = None
            print("ℹ  FIDE plugin not found – FIDE features disabled.")

        self._register_builtin_strategies()
        self._register_builtin_calculators()

        self.plugin_loader = PluginLoader(
            self.strategy_registry, self.calculator_registry, self.repository
        )

        self.service = TournamentService(
            self.repository, self.strategy_registry, self.calculator_registry
        )

        self.plugin_loader.discover_and_load_plugins()
        
        
        self._create_default_admin()

    def _register_builtin_strategies(self):
        self.strategy_registry.register(RoundRobinStrategy(self.repository))
        self.strategy_registry.register(SingleEliminationStrategy(self.repository))
        self.strategy_registry.register(SwissStrategy(self.repository))
        self.strategy_registry.register(FreeForAllStrategy(self.repository))
        if FIDE_AVAILABLE and self.fide_db is not None:
            self.strategy_registry.register(
                ChessFideSwissStrategy(self.repository, self.fide_db)
            )

    def _register_builtin_calculators(self):
        self.calculator_registry.register(StandardPointsCalculator())
        self.calculator_registry.register(ThreePointsCalculator())
        self.calculator_registry.register(RankingPointsCalculator())
        self.calculator_registry.register(PercentagePointsCalculator())
        if FIDE_AVAILABLE:
            self.calculator_registry.register(ChessFidePointsCalculator())
    
    def _create_default_admin(self):
        """Create default admin account if none exists."""
        try:
            admin = self.user_repository.get_admin_by_username("admin")
            if not admin:
                print("Creating default admin account...")
                self.user_mgmt_service.register_admin(
                    username="admin",
                    password="admin123",
                    email="admin@tournament.local"
                )
                print(" Default admin created - Username: admin, Password: admin123")
                print(" IMPORTANT: Change the admin password immediately!")
        except Exception as e:
            print(f"Warning: Could not create default admin: {e}")



app_service = TournamentWebApp()


auth_blueprint = init_auth_routes(
    app_service.auth_service,
    app_service.authz_service,
    app_service.user_mgmt_service,
    app_service.auth_middleware
)
app.register_blueprint(auth_blueprint)


@app.route("/")
def index():
    """Main page."""
    return render_template("main.html")


@app.route("/login")
def login_page():
    """Login page."""
    return render_template("login.html")






@app.route("/api/players", methods=["GET"])
@app_service.auth_middleware.require_staff_or_admin
def get_players():
    """Get all players (staff/admin only)."""
    try:
        players = app_service.service.list_players()
        return jsonify(
            [
                {
                    "id": p.id,
                    "name": p.name,
                    "short_id": p.id[:8],
                    "date_of_birth": p.date_of_birth,
                }
                for p in players
            ]
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/players", methods=["POST"])
@app_service.auth_middleware.require_staff_or_admin
def create_player():
    """Create a new player (staff/admin only)."""
    try:
        data = request.get_json()
        name = data.get("name", "").strip()
        date_of_birth = data.get("date_of_birth", "").strip() or None
        rating_raw = data.get("rating", "")
        title_raw = data.get("title", "")
        federation_raw = data.get("federation", "")
        club = data.get("club", "")
        fide_id_raw = data.get("fide_id", "")
        if not name:
            return jsonify({"error": "Name is required"}), 400

        try:
            rating = _parse_fide_rating(rating_raw)
            title = _normalize_fide_title(title_raw)
            federation = _normalize_fide_federation(federation_raw)
            fide_id = _normalize_fide_id(fide_id_raw)
            if date_of_birth:
                date_of_birth = _validate_fide_date(date_of_birth, "date_of_birth")
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

        if fide_id:
            existing = app_service.fide_db.get_player_by_fide_id(fide_id)
            if existing:
                return jsonify({"error": "FIDE ID already assigned to another player"}), 409

        player_id = app_service.service.create_player(name, date_of_birth=date_of_birth)
        if date_of_birth:
            try:
                app_service.user_mgmt_service.register_player(
                    player_id, name, date_of_birth
                )
            except ValueError as e:
                app_service.repository.delete_player(player_id)
                return jsonify({"error": str(e)}), 409
        has_fide_fields = any(
            [
                date_of_birth,
                rating_raw not in (None, ""),
                str(title_raw).strip(),
                str(federation_raw).strip(),
                str(club).strip(),
                str(fide_id_raw).strip(),
            ]
        )
        if has_fide_fields:
            try:
                app_service.fide_db.register_player(
                    player_id, name, rating, date_of_birth or "", federation, title, club, fide_id
                )
            except ValueError as e:
                app_service.repository.delete_player(player_id)
                return jsonify({"error": str(e)}), 409
        return (
            jsonify(
                {
                    "id": player_id,
                    "name": name,
                    "date_of_birth": date_of_birth,
                }
            ),
            201,
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 409
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/players/<player_id>", methods=["DELETE"])
@app_service.auth_middleware.require_staff_or_admin
def delete_player(player_id):
    """Delete a player (staff/admin only)."""
    try:
        app_service.repository.delete_player(player_id)
        return jsonify({"message": "Player deleted"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500






@app.route("/api/tournaments", methods=["GET"])
@app_service.auth_middleware.require_auth
def get_tournaments():
    """Get all tournaments (authenticated users)."""
    try:
        user = app_service.auth_middleware.get_current_user()
        
        with app_service.repository._get_connection() as conn:
            rows = conn.execute(
                "SELECT id, name, created_at FROM tournaments ORDER BY created_at DESC"
            ).fetchall()
            
            tournaments = [
                {
                    "id": r["id"],
                    "name": r["name"],
                    "short_id": r["id"][:8],
                    "created_at": r["created_at"],
                }
                for r in rows
            ]
            
            
            if user.role == UserRole.PLAYER:
                player_tournaments = []
                for t in tournaments:
                    players = app_service.repository.get_tournament_players(t["id"])
                    if any(p["player_id"] == user.player_id for p in players):
                        player_tournaments.append(t)
                return jsonify(player_tournaments)
            
            return jsonify(tournaments)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tournaments", methods=["POST"])
@app_service.auth_middleware.require_permission(
    lambda user: app_service.authz_service.can_create_tournament(user)
)
def create_tournament():
    """Create a new tournament (staff/admin only)."""
    try:
        data = request.get_json()
        name = data.get("name", "").strip()
        if not name:
            return jsonify({"error": "Name is required"}), 400

        tournament_id = app_service.service.create_tournament(name)
        return jsonify({"id": tournament_id, "name": name}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tournaments/<tournament_id>", methods=["GET"])
@app_service.auth_middleware.require_auth
def get_tournament(tournament_id):
    """Get tournament details."""
    try:
        user = app_service.auth_middleware.get_current_user()
        
        with app_service.repository._get_connection() as conn:
            tournament = conn.execute(
                "SELECT * FROM tournaments WHERE id = ?", (tournament_id,)
            ).fetchone()

            if not tournament:
                return jsonify({"error": "Tournament not found"}), 404

            players = app_service.repository.get_tournament_players(tournament_id)
            
            
            if user.role == UserRole.PLAYER:
                if not any(p["player_id"] == user.player_id for p in players):
                    return jsonify({"error": "Access denied"}), 403

            
            is_fide = False
            if FIDE_AVAILABLE and app_service.fide_db is not None:
                try:
                    fide_t = app_service.fide_db.get_tournament(tournament_id)
                    is_fide = fide_t is not None
                except Exception:
                    is_fide = False

            return jsonify(
                {
                    "id": tournament["id"],
                    "name": tournament["name"],
                    "is_fide": is_fide,
                    "type": "fide" if is_fide else "standard",
                    "players": [
                        {
                            "player_id": p["player_id"],
                            "name": p["name"],
                            "able_to_play": p.get("able_to_play", 1),
                            "short_id": p["player_id"][:8],
                        }
                        for p in players
                    ],
                }
            )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tournaments/<tournament_id>/players", methods=["POST"])
@app_service.auth_middleware.require_permission(
    lambda user: app_service.authz_service.can_manage_players(user)
)
def add_players_to_tournament(tournament_id):
    """Add players to tournament (staff/admin only)."""
    try:
        data = request.get_json()
        player_ids = data.get("player_ids", [])

        if not player_ids:
            return jsonify({"error": "No players specified"}), 400

        for player_id in player_ids:
            app_service.service.add_player_to_tournament(tournament_id, player_id)

        return jsonify({"message": f"Added {len(player_ids)} players"}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tournaments/<tournament_id>/eliminate-player", methods=["POST"])
@app_service.auth_middleware.require_staff_or_admin
def eliminate_player(tournament_id):
    """Eliminate a player from tournament."""
    try:
        data = request.get_json()
        player_id = data.get("player_id")

        if not player_id:
            return jsonify({"error": "player_id required"}), 400

        app_service.repository.eliminate_player(tournament_id, player_id)
        return jsonify({"message": "Player eliminated"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tournaments/<tournament_id>/activate-player", methods=["POST"])
@app_service.auth_middleware.require_staff_or_admin
def activate_player(tournament_id):
    """Reactivate a player in tournament."""
    try:
        data = request.get_json()
        player_id = data.get("player_id")

        if not player_id:
            return jsonify({"error": "player_id required"}), 400

        app_service.repository.activate_player(tournament_id, player_id)
        return jsonify({"message": "Player activated"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500






@app.route("/api/tournaments/<tournament_id>/rounds", methods=["POST"])
@app_service.auth_middleware.require_permission(
    lambda user: app_service.authz_service.can_create_round(user)
)
def create_round(tournament_id):
    """Create a new round (staff/admin only)."""
    try:
        data = request.get_json()
        print(f"DEBUG: create_round received data: {data}")
        
        round_type = data.get("round_type") or data.get("strategy", "roundrobin")
        players_per_match = data.get("players_per_match", 2)
        force_create = data.get("force", False)

        
        
        is_fide_tournament = False
        if FIDE_AVAILABLE and app_service.fide_db is not None:
            try:
                fide_t = app_service.fide_db.get_tournament(tournament_id)
                is_fide_tournament = fide_t is not None
            except Exception:
                pass

        if is_fide_tournament and round_type != "chess_fide_swiss":
            return jsonify({
                "error": (
                    f"This is a FIDE Chess tournament. Only the 'chess_fide_swiss' "
                    f"strategy is permitted. Received: '{round_type}'."
                )
            }), 400

        if not is_fide_tournament and round_type == "chess_fide_swiss":
            return jsonify({
                "error": (
                    "The 'chess_fide_swiss' strategy is only available for FIDE Chess "
                    "tournaments. Use the FIDE section to create a FIDE tournament first."
                )
            }), 400
        

        config = RoundConfig(
            tournament_id=tournament_id,
            round_type=round_type,
            players_per_match=players_per_match,
            force_create=force_create,
        )

        result = app_service.service.create_round(config)

        return jsonify(
            {
                "round_id": result["round_id"],
                "ordinal": result["ordinal"],
                "matches_count": len(result["matches"]),
                "waiting_players": result.get("waiting_players", []),
            }
        ), 201
    except ValueError as e:
        if "pending matches" in str(e).lower():
            return jsonify({"error": str(e), "code": "PENDING_MATCHES"}), 409
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tournaments/<tournament_id>/rounds", methods=["GET"])
@app_service.auth_middleware.require_auth
def get_tournament_rounds(tournament_id):
    """Get all rounds for a tournament."""
    try:
        with app_service.repository._get_connection() as conn:
            rows = conn.execute(
                """
                SELECT id, round_type, ordinal, created_at 
                FROM rounds 
                WHERE tournament_id = ? 
                ORDER BY ordinal
                """,
                (tournament_id,),
            ).fetchall()

            return jsonify(
                [
                    {
                        "id": r["id"],
                        "round_type": r["round_type"],
                        "ordinal": r["ordinal"],
                        "short_id": r["id"][:8],
                        "created_at": r["created_at"],
                    }
                    for r in rows
                ]
            )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rounds/<round_id>/matches", methods=["GET"])
@app_service.auth_middleware.require_auth
def get_round_matches(round_id):
    """Get matches for a round."""
    try:
        user = app_service.auth_middleware.get_current_user()
        matches = app_service.repository.list_matches_for_round(round_id)
        
        
        if user.status == UserStatus.SHADOW_BANNED:
            return jsonify([])
        
        
        if user.role == UserRole.PLAYER:
            matches = [m for m in matches if user.player_id in m.player_ids]

        match_data = []
        for match in matches:
            
            players = []
            player_names = []
            for pid in match.player_ids:
                player = app_service.repository.get_player(pid)
                if player:
                    players.append({"id": pid, "name": player.name, "short_id": pid[:8]})
                    player_names.append(player.name)
                else:
                    players.append({"id": pid, "name": pid[:8], "short_id": pid[:8]})
                    player_names.append(pid[:8])

            
            winner_names = []
            if match.winner_ids:
                for wid in match.winner_ids:
                    wp = app_service.repository.get_player(wid)
                    if wp:
                        winner_names.append(wp.name)

            
            if match.auto_bye:
                status = "bye"
            elif match.result in ("draw",):
                status = "draw"
            elif match.result in ("complete", "auto") or match.winner_ids:
                status = "completed"
            else:
                status = "pending"

            
            colors = match.colors or []
            white_player = None
            black_player = None
            white_id = None
            black_id = None
            for idx_c, pid in enumerate(match.player_ids):
                color = colors[idx_c] if idx_c < len(colors) else None
                p_obj = next((p for p in players if p["id"] == pid), None)
                pname = p_obj["name"] if p_obj else pid[:8]
                if color == "white":
                    white_player = pname
                    white_id = pid
                elif color == "black":
                    black_player = pname
                    black_id = pid

            match_data.append(
                {
                    "id": match.id,
                    "short_id": match.id[:8],
                    
                    "player_ids": match.player_ids,
                    "players": players,
                    "result": match.result,
                    "winner_ids": match.winner_ids or [],
                    "auto_bye": match.auto_bye,
                    "players_per_match": match.players_per_match,
                    
                    "board_no": match.board_no,
                    "white_player": white_player,
                    "white_id": white_id,
                    "black_player": black_player,
                    "black_id": black_id,
                    
                    "player_names": player_names,
                    "winner_names": winner_names,
                    "status": status,
                }
            )

        return jsonify(match_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500






@app.route("/api/matches/<match_id>/result", methods=["POST"])
@app_service.auth_middleware.require_permission(
    lambda user: app_service.authz_service.can_record_result(user)
)
def record_match_result(match_id):
    """Record match result (staff/admin only)."""
    try:
        data = request.get_json()

        
        result_type = data.get("result") or data.get("result_type", "")

        
        winner_ids = data.get("winner_ids") or []
        if not winner_ids:
            single_winner = data.get("winner_id")
            if single_winner:
                winner_ids = [single_winner]

        rankings = data.get("rankings", {})
        calculator = data.get("calculator", None)

        is_draw = result_type in ("draw",)

        
        if result_type == "rankings" and rankings and not winner_ids:
            winner_ids = [pid for pid, rank in rankings.items() if rank == 1]

        match_result = MatchResult(
            match_id=match_id,
            winner_ids=winner_ids if not is_draw else [],
            rankings=rankings,
            is_draw=is_draw,
        )

        app_service.service.record_match_result(match_id, match_result, calculator)

        return jsonify({"message": "Result recorded successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500






@app.route("/api/tournaments/<tournament_id>/standings", methods=["GET"])
@app_service.auth_middleware.require_auth
def get_standings(tournament_id):
    """Get tournament standings."""
    try:
        user = app_service.auth_middleware.get_current_user()
        
        
        if user.status == UserStatus.SHADOW_BANNED:
            return jsonify([])
        
        standings = app_service.service.get_standings(tournament_id)
        
        
        if user.role == UserRole.PLAYER:
            players = app_service.repository.get_tournament_players(tournament_id)
            if not any(p["player_id"] == user.player_id for p in players):
                return jsonify({"error": "Access denied"}), 403

        return jsonify(standings)
    except Exception as e:
        return jsonify({"error": str(e)}), 500






@app.route("/api/player/my-pairings/<tournament_id>", methods=["GET"])
@app_service.auth_middleware.require_role(UserRole.PLAYER)
def get_player_pairings(tournament_id):
    """Get player's pairings in latest round (players only)."""
    try:
        user = app_service.auth_middleware.get_current_user()
        
        
        if user.status == UserStatus.SHADOW_BANNED:
            return jsonify({"message": "No pairings found", "pairings": []}), 200
        
        
        with app_service.repository._get_connection() as conn:
            latest_round = conn.execute("""
                SELECT id, ordinal FROM rounds 
                WHERE tournament_id = ? 
                ORDER BY ordinal DESC LIMIT 1
            """, (tournament_id,)).fetchone()
            
            if not latest_round:
                return jsonify({"message": "No rounds yet", "pairings": []}), 200
            
            
            matches = app_service.repository.list_matches_for_round(latest_round["id"])
            player_matches = [m for m in matches if user.player_id in m.player_ids]
            
            
            pairings = []
            for match in player_matches:
                opponents = []
                for pid in match.player_ids:
                    if pid != user.player_id:
                        player = app_service.repository.get_player(pid)
                        if player:
                            opponents.append({"id": pid, "name": player.name})
                
                pairings.append({
                    "match_id": match.id,
                    "round": latest_round["ordinal"],
                    "opponents": opponents,
                    "result": match.result,
                    "is_bye": match.auto_bye
                })
            
            return jsonify({
                "tournament_id": tournament_id,
                "round": latest_round["ordinal"],
                "pairings": pairings
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500






@app.route("/api/strategies", methods=["GET"])
@app_service.auth_middleware.require_staff_or_admin
def list_strategies():
    """List available matchmaking strategies."""
    try:
        strategies = app_service.service.list_available_strategies()
        
        return jsonify(strategies)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/calculators", methods=["GET"])
@app_service.auth_middleware.require_staff_or_admin
def list_calculators():
    """List available points calculators."""
    try:
        calculators = app_service.service.list_available_calculators()
        
        return jsonify(calculators)
    except Exception as e:
        return jsonify({"error": str(e)}), 500






def _validate_fide_date(value: str, field_name: str) -> str:
    if not value:
        return ""
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        raise ValueError(f"Invalid {field_name} format (YYYY-MM-DD required)")
    return value


def _normalize_fide_federation(value: str) -> str:
    if not value:
        return ""
    fed = value.strip().upper()
    if len(fed) != 3:
        raise ValueError("Federation must be a 3-letter code")
    return fed


def _normalize_fide_title(value: str) -> str:
    title = value.strip().upper() if value else ""
    if title and title not in FIDE_TITLES:
        raise ValueError("Invalid FIDE title")
    return title


def _parse_fide_rating(value) -> int:
    if value in (None, ""):
        return 0
    try:
        rating = int(value)
    except (TypeError, ValueError):
        raise ValueError("Rating must be an integer")
    if rating < 0:
        raise ValueError("Rating cannot be negative")
    return rating


def _normalize_fide_id(value: str) -> str:
    if not value:
        return ""
    fid = value.strip()
    if not fid.isdigit():
        raise ValueError("FIDE ID must be numeric")
    return fid


def _normalize_fide_categories(value) -> list[str]:
    if not value:
        return ["Open"]
    if isinstance(value, str):
        cats = [c.strip() for c in value.split(",") if c.strip()]
    else:
        cats = [str(c).strip() for c in value if str(c).strip()]
    for c in cats:
        if c not in FIDE_CATEGORIES:
            raise ValueError(f"Invalid category: {c}")
    return list(dict.fromkeys(cats)) or ["Open"]


def _require_fide():
    """Return a 503 response if the FIDE plugin is not installed."""
    if not FIDE_AVAILABLE:
        from flask import jsonify as _jsonify
        return _jsonify({"error": "FIDE plugin is not installed. Please add plugins/chess_fide_swiss.py."}), 503
    return None


@app.route("/api/fide/tournaments", methods=["POST"])
@app_service.auth_middleware.require_staff_or_admin
def fide_create_tournament():
    """Create a FIDE chess tournament."""
    err = _require_fide()
    if err:
        return err
    try:
        data = request.get_json()
        name = data.get("name", "").strip()
        arbiter = data.get("arbiter", "").strip()
        chief_arbiter = data.get("chief_arbiter", "").strip()
        time_control = data.get("time_control", "").strip()
        tournament_date = data.get("tournament_date", "").strip()
        max_rounds = data.get("max_rounds", 7)
        venue = data.get("venue", "").strip()
        categories = data.get("categories", [])

        if not name:
            return jsonify({"error": "Name required"}), 400
        if not tournament_date:
            return jsonify({"error": "Tournament date required"}), 400

        tournament_date = _validate_fide_date(tournament_date, "tournament_date")
        try:
            max_rounds = int(max_rounds)
        except (TypeError, ValueError):
            return jsonify({"error": "Max rounds must be an integer"}), 400
        if max_rounds < 1 or max_rounds > 30:
            return jsonify({"error": "Max rounds must be between 1 and 30"}), 400

        try:
            categories = _normalize_fide_categories(categories)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

        
        tid = app_service.service.create_tournament(name)

        
        app_service.fide_db.create_tournament(
            tid,
            name,
            tournament_date,
            max_rounds,
            venue,
            arbiter,
            chief_arbiter or arbiter,
            time_control or "Standard",
            categories,
        )

        return jsonify({"id": tid, "name": name}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fide/tournaments/<tid>/players", methods=["POST"])
@app_service.auth_middleware.require_staff_or_admin
def fide_add_player(tid):
    """Add player to FIDE tournament."""
    try:
        data = request.get_json()
        player = None
        
        
        player_id = data.get("player_id", "")
        if player_id:
            player = app_service.repository.get_player(player_id)
            if not player:
                return jsonify({"error": "Player not found"}), 404
            name = player.name
            pid = player.id
        else:
            
            name = data.get("name", "").strip()
            if not name:
                return jsonify({"error": "Name required"}), 400
            pid = app_service.service.create_player(name)

        rating = data.get("rating", 0)
        title = data.get("title", "")
        federation = data.get("federation", "")
        fide_id = data.get("fide_id", "")
        club = data.get("club", "")
        dob = data.get("dob") or data.get("date_of_birth") or ""
        if player and not dob:
            dob = player.date_of_birth or ""

        try:
            rating = _parse_fide_rating(rating)
            title = _normalize_fide_title(title)
            federation = _normalize_fide_federation(federation)
            fide_id = _normalize_fide_id(fide_id)
            dob = _validate_fide_date(dob, "date_of_birth") if dob else ""
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

        tournament = app_service.fide_db.get_tournament(tid)
        if not tournament:
            return jsonify({"error": "Tournament not found"}), 404
        categories = tournament.get("categories") or ["Open"]
        if dob:
            category = app_service.fide_db.auto_assign_category(
                dob, tournament["tournament_date"], categories
            )
        else:
            category = "Open" if "Open" in categories else categories[0]
        if category not in categories:
            category = categories[0]

        
        app_service.service.add_player_to_tournament(tid, pid)

        
        app_service.fide_db.register_player(
            pid, name, rating, dob, federation, title, club, fide_id
        )
        app_service.fide_db.add_player_to_tournament(tid, pid, category)

        return jsonify({"player_id": pid, "name": name}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fide/tournaments/<tid>/players/<pid>", methods=["DELETE"])
@app_service.auth_middleware.require_staff_or_admin
def fide_remove_player(tid, pid):
    """Remove player from FIDE tournament."""
    try:
        
        app_service.fide_db.remove_player(tid, pid)
        
        
        
        
        
        
        
        
        
        
        with app_service.repository._get_connection() as conn:
            conn.execute(
                "DELETE FROM tournament_players WHERE tournament_id = ? AND player_id = ?",
                (tid, pid)
            )
            conn.commit()

        return jsonify({"message": "Player removed"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fide/tournaments/<tid>/rounds", methods=["POST"])
@app_service.auth_middleware.require_staff_or_admin
def fide_create_round(tid):
    """Create FIDE Swiss round."""
    try:
        data = request.get_json() or {}
        force_create = data.get("force", False)

        config = RoundConfig(
            tournament_id=tid,
            round_type="fide_swiss",
            players_per_match=2,
            force_create=force_create
        )

        result = app_service.service.create_round(config)

        return jsonify({
            "round_id": result["round_id"],
            "ordinal": result["ordinal"],
            "matches_count": len(result["matches"])
        }), 201
    except ValueError as e:
        if "pending matches" in str(e).lower():
            return jsonify({"error": str(e), "code": "PENDING_MATCHES"}), 409
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fide/tournaments/<tid>/rounds/<int:round_num>/pairings", methods=["GET"])
@app_service.auth_middleware.require_auth
def fide_get_pairings(tid, round_num):
    """Get FIDE round pairings."""
    try:
        pairings = app_service.fide_reports.round_pairings(tid, round_num)
        return jsonify(pairings)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fide/tournaments/<tid>/standings", methods=["GET"])
@app_service.auth_middleware.require_auth
def fide_get_standings(tid):
    """Get FIDE tournament standings with tiebreaks."""
    try:
        category = request.args.get("category")
        standings = app_service.fide_reports.standings(tid, category_filter=category)
        return jsonify(standings)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fide/matches/<mid>/result", methods=["POST"])
@app_service.auth_middleware.require_staff_or_admin
def fide_record_result(mid):
    """Record FIDE chess match result."""
    try:
        data = request.get_json()
        result_str = data.get("result", "").strip()

        
        match = app_service.repository.get_match(mid)
        if not match:
            return jsonify({"error": "Match not found"}), 404

        if len(match.player_ids) != 2:
            return jsonify({"error": "FIDE matches must have 2 players"}), 400

        white_id, black_id = match.player_ids[0], match.player_ids[1]
        winner_ids = []
        is_draw = False

        
        if result_str in ("1-0", "white"):
            winner_ids = [white_id]
        elif result_str in ("0-1", "black"):
            winner_ids = [black_id]
        elif result_str in ("0.5-0.5", "draw", "1/2-1/2"):
            is_draw = True
        else:
            return jsonify({"error": f"Invalid result: {result_str}"}), 400

        match_result = MatchResult(
            match_id=mid,
            winner_ids=winner_ids,
            is_draw=is_draw,
            rankings={},
        )

        app_service.service.record_match_result(
            mid, match_result, calculator_name="fide_standard"
        )

        
        tid = match.tournament_id
        round_id = match.round_id
        
        fide_rounds = app_service.fide_db.get_fide_rounds(tid)
        round_ordinal = None
        for fr in fide_rounds:
            if fr["round_id"] == round_id:
                round_ordinal = fr["round_ordinal"]
                break

        if round_ordinal is not None:
            if is_draw:
                app_service.fide_db.update_color_result(tid, round_ordinal, white_id, "0.5")
                app_service.fide_db.update_color_result(tid, round_ordinal, black_id, "0.5")
            elif white_id in winner_ids:
                app_service.fide_db.update_color_result(tid, round_ordinal, white_id, "1")
                app_service.fide_db.update_color_result(tid, round_ordinal, black_id, "0")
            else:
                app_service.fide_db.update_color_result(tid, round_ordinal, white_id, "0")
                app_service.fide_db.update_color_result(tid, round_ordinal, black_id, "1")

        return jsonify({"message": "Result recorded", "result": result_str})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fide/tournaments/<tid>/export", methods=["GET"])
@app_service.auth_middleware.require_auth
def fide_export_tournament_excel(tid):
    """Export FIDE tournament data to Excel."""
    try:
        t = app_service.fide_db.get_tournament(tid)
        if not t:
            return jsonify({"error": "Tournament not found"}), 404

        round_param = request.args.get('round', 'all')
        
        
        def style_worksheet(ws):
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="363636", end_color="363636", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            
            for column_cells in ws.columns:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 4
                
                
                for cell in column_cells:
                    cell.border = thin_border
                    if cell.row > 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            
            if round_param != 'all':
                try:
                    r_num = int(round_param)
                    pairings = app_service.fide_reports.round_pairings(tid, r_num)
                    
                    data = []
                    for p in pairings:
                        w_info = f"{p.get('white_name', '')} ({p.get('white_rating', 0)})"
                        b_info = f"{p.get('black_name', '')} ({p.get('black_rating', 0)})"
                        res = f"{p.get('white_result','')} - {p.get('black_result','')}" if p.get('white_result') else "   -   "
                        
                        data.append({
                            "Board": p["board"],
                            "White": w_info,
                            "Result": res,
                            "Black": b_info
                        })
                    
                    df = pd.DataFrame(data)
                    sheet_name = f"Round {r_num} Pairings"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    style_worksheet(writer.sheets[sheet_name])
                    
                    ws = writer.sheets[sheet_name]
                    for row in ws.iter_rows(min_row=2):
                        row[0].alignment = Alignment(horizontal="center")
                        row[2].alignment = Alignment(horizontal="center")

                except ValueError:
                    return jsonify({"error": "Invalid round number"}), 400

            
            else:
                
                players = app_service.fide_db.get_tournament_players(tid)
                df_players = pd.DataFrame(players)
                if not df_players.empty:
                    df_players = df_players.rename(columns={
                        "name": "Name", "rating": "Rating", "title": "Title", 
                        "federation": "Fed", "fide_id": "FIDE ID", "starting_rank": "Rank"
                    })
                    cols = ["Rank", "Name", "Title", "Rating", "Fed", "FIDE ID"]
                    df_players = df_players[[c for c in cols if c in df_players.columns]]
                
                df_players.to_excel(writer, sheet_name='Players', index=False)
                style_worksheet(writer.sheets['Players'])

                
                standings = app_service.fide_reports.standings(tid)
                standing_data = []
                for s in standings:
                    row = {
                        "Rank": s["rank"],
                        "Name": s["name"],
                        "Points": s["points"],
                        "Title": s.get("title", ""),
                        "Fed": s.get("federation", ""),
                        "Rating": s.get("rating", 0),
                        "BH": s.get("tiebreaks", {}).get("buchholz", 0),
                        "SB": s.get("tiebreaks", {}).get("sonneborn_berger", 0),
                        "Wins": s.get("num_wins", 0)
                    }
                    standing_data.append(row)
                
                df_standings = pd.DataFrame(standing_data)
                df_standings.to_excel(writer, sheet_name='Standings', index=False)
                style_worksheet(writer.sheets['Standings'])

                
                rounds = app_service.fide_db.get_fide_rounds(tid)
                all_pairings = []
                for r in rounds:
                    pairings = app_service.fide_reports.round_pairings(tid, r["round_ordinal"])
                    for p in pairings:
                        all_pairings.append({
                            "Round": r["round_ordinal"],
                            "Board": p["board"],
                            "White": p.get("white_name", ""),
                            "Result": f"{p.get('white_result','')} - {p.get('black_result','')}" if p.get('white_result') else "?-?",
                            "Black": p.get("black_name", ""),
                        })
                
                df_pairings = pd.DataFrame(all_pairings)
                df_pairings.to_excel(writer, sheet_name='All Pairings', index=False)
                style_worksheet(writer.sheets['All Pairings'])

        output.seek(0)
        
        safe_name = t['name'].replace(' ', '_')
        suffix = f"Round_{round_param}" if round_param != 'all' else "Full_Export"
        filename = f"{safe_name}_{suffix}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fide/tournaments/<tid>/title-norms/<player_id>", methods=["GET"])
@app_service.auth_middleware.require_auth
def fide_check_title_norms(tid, player_id):
    """Check if player achieved title norms."""
    try:
        norms = app_service.title_norm_checker.check_title_norms(tid, player_id)
        return jsonify(norms)
    except Exception as e:
        return jsonify({"error": str(e)}), 500






@app.route("/api/admin/dashboard/stats", methods=["GET"])
@app_service.auth_middleware.require_admin
def admin_dashboard_stats():
    """Get admin dashboard statistics."""
    try:
        with app_service.repository._get_connection() as conn:
            
            tournaments_count = conn.execute("SELECT COUNT(*) as cnt FROM tournaments").fetchone()["cnt"]
            
            
            players_count = conn.execute("SELECT COUNT(*) as cnt FROM players").fetchone()["cnt"]
            
            
            matches_count = conn.execute("SELECT COUNT(*) as cnt FROM matches").fetchone()["cnt"]
            
            
            admin_count = conn.execute("SELECT COUNT(*) as cnt FROM admin_users").fetchone()["cnt"]
            staff_count = conn.execute("SELECT COUNT(*) as cnt FROM staff_users WHERE status = 'active'").fetchone()["cnt"]
            pending_staff = conn.execute("SELECT COUNT(*) as cnt FROM staff_users WHERE status = 'pending'").fetchone()["cnt"]
            player_users_count = conn.execute("SELECT COUNT(*) as cnt FROM player_users").fetchone()["cnt"]
            
            return jsonify({
                "tournaments": tournaments_count,
                "players": players_count,
                "matches": matches_count,
                "admins": admin_count,
                "staff": staff_count,
                "pending_staff": pending_staff,
                "player_accounts": player_users_count
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/audit-log", methods=["GET"])
@app_service.auth_middleware.require_admin
def get_audit_log():
    """Get audit log entries (admin only)."""
    try:
        limit = request.args.get('limit', 100, type=int)
        
        with app_service.user_repository._get_connection() as conn:
            rows = conn.execute("""
                SELECT al.*, au.username as admin_username
                FROM audit_log al
                LEFT JOIN admin_users au ON al.admin_id = au.id
                ORDER BY al.timestamp DESC
                LIMIT ?
            """, (limit,)).fetchall()
            
            return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/users", methods=["GET"])
@app_service.auth_middleware.require_admin
def list_all_users():
    """List all users (admin only)."""
    try:
        with app_service.user_repository._get_connection() as conn:
            
            admins = conn.execute("SELECT id, username, email, status, created_at, last_login FROM admin_users").fetchall()
            admin_list = [{"role": "admin", **dict(r)} for r in admins]
            
            
            staff = conn.execute("SELECT id, username, email, status, created_at, last_login, approved_by FROM staff_users").fetchall()
            staff_list = [{"role": "staff", **dict(r)} for r in staff]
            
            
            players = conn.execute("SELECT id, player_id, name, date_of_birth, status, created_at, last_login FROM player_users").fetchall()
            player_list = [{"role": "player", **dict(r)} for r in players]
            
            return jsonify({
                "admins": admin_list,
                "staff": staff_list,
                "players": player_list
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":

    
    app.run(debug=True, host="0.0.0.0", port=5000)